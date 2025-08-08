#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FILE: whatsapp_outgoing_filter.py

INPUT FILES:
- WhatsApp CSV (ground truth direction):
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM/Arjun Digital Identity/WhatsApp - 471 chat sessions.csv
  Description: Original WhatsApp export with explicit direction in column "Type" (Outgoing/Incoming/Notification).
  Required columns: ["Chat Session", "Message Date", "Sent Date", "Type", "Sender ID", "Sender Name", "Status", "Replying to", "Text", "Attachment", "Attachment type", "Attachment info"]

- Training JSONL to filter:
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_training_dataset.jsonl
  Description: Instruction-style records with fields like: instruction, input, output, source, persona_weight, chunk_id, timestamp, scores

- Validation JSONL to filter:
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_validation_dataset.jsonl
  Description: Same schema as training JSONL.

OUTPUT FILES:
- Filtered Training JSONL (whatsapp outgoing only; others unchanged):
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_training_dataset.whatsapp_outgoing_only.jsonl
  Format: line-delimited JSON objects; WhatsApp entries retained only if matched to an Outgoing CSV row; non-WhatsApp entries retained as-is.

- Filtered Validation JSONL (whatsapp outgoing only; others unchanged):
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_validation_dataset.whatsapp_outgoing_only.jsonl
  Format: same as above.

- Audit Report (XLSX):
  /Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/whatsapp_filter_audit.xlsx
  Sheets:
    1) summary: counts per file (kept, dropped_incoming_or_notification, dropped_no_match)
    2) kept_matches: JSONL record + matched CSV row (Type=Outgoing)
    3) dropped_matches: JSONL record + matched CSV row (Type not Outgoing)
    4) no_match: JSONL record with no reliable CSV match

Version history:
- v1.0 (2025-08-07): Initial implementation with strict filtering and LA timezone handling

Notes and methodology (readable for a 10th grader):
- We treat the CSV as the source of truth because it explicitly says whether a message is Outgoing (you wrote it) or Incoming (someone else wrote it).
- For every JSONL record with source=="whatsapp", we try to find the same message in the CSV by:
  1) Comparing the message text (after simple cleanup) and
  2) Making sure the times are close (within 5 minutes). The CSV time is in Los Angeles time; we convert it to UTC to compare with the JSONL time.
- If that fails, we try a time-only match within 90 seconds. If there's exactly one outgoing message around that time, we accept it.
- If we cannot reliably match a JSONL message to an Outgoing CSV row, we drop it (strict mode). Non-WhatsApp JSONL rows are kept unchanged.
- We write an Excel report so you can see what we kept and what we dropped.

Missing data handling:
- If CSV "Sent Date" is missing, that row is ignored for matching.
- If JSONL timestamp is missing or invalid, that JSONL row cannot be matched and will be dropped when source=="whatsapp" (strict), and recorded in the audit.
- This script does not impute country data; not applicable for this dataset.

"""

import csv
import json
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Constants and paths
LA_TZ = ZoneInfo("America/Los_Angeles")
UTC_TZ = ZoneInfo("UTC")

CSV_PATH = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM/Arjun Digital Identity/WhatsApp - 471 chat sessions.csv")
TRAIN_JSONL = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_training_dataset.jsonl")
VAL_JSONL = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_validation_dataset.jsonl")
OUT_TRAIN_JSONL = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_training_dataset.whatsapp_outgoing_only.jsonl")
OUT_VAL_JSONL = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/persona_validation_dataset.whatsapp_outgoing_only.jsonl")
AUDIT_XLSX = Path("/Users/macbook2024/Dropbox/AAA Backup/A Working/Arjun LLM Writing/whatsapp_filter_audit.xlsx")

# Matching tolerances
TEXT_TIME_WINDOW = timedelta(minutes=5)
TIME_ONLY_WINDOW = timedelta(seconds=90)


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    t = s.strip()
    # Replace smart quotes and normalize whitespace
    replacements = {
        "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"',
        "\u2013": "-", "\u2014": "-",
        "\u00a0": " ",
    }
    for a, b in replacements.items():
        t = t.replace(a, b)
    # Remove common reply arrow prefixes if they leaked into text
    # CSV has a separate "Replying to" column; JSONL output shouldn't include it.
    if t.startswith("\u279c") or t.startswith("➜"):
        # Remove leading arrow and any spaces
        t = t.lstrip("\u279c➜ ")
    # Collapse whitespace
    t = " ".join(t.split())
    return t


def parse_csv_datetime(sent_date_str: str) -> Optional[datetime]:
    """Parse CSV "Sent Date" (local LA time, no timezone) and return aware UTC datetime."""
    if not sent_date_str:
        return None
    try:
        # Example: 2022-08-05 10:08:32
        dt_naive = datetime.strptime(sent_date_str, "%Y-%m-%d %H:%M:%S")
        dt_local = dt_naive.replace(tzinfo=LA_TZ)
        return dt_local.astimezone(UTC_TZ)
    except Exception:
        return None


def parse_jsonl_timestamp(ts: str) -> Optional[datetime]:
    if not ts:
        return None
    try:
        # Python 3.11 handles fromisoformat with timezone like "+00:00"
        return datetime.fromisoformat(ts).astimezone(UTC_TZ)
    except Exception:
        return None


@dataclass
class CsvRow:
    chat_session: str
    sent_dt_utc: datetime
    type: str
    text_norm: str
    raw_text: str


def load_outgoing_csv(csv_path: Path) -> Tuple[List[CsvRow], List[datetime]]:
    outgoing_rows: List[CsvRow] = []
    time_index: List[datetime] = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            type_val = (row.get("Type") or "").strip()
            if type_val != "Outgoing":
                continue
            sent_date = row.get("Sent Date") or ""
            sent_dt_utc = parse_csv_datetime(sent_date)
            if not sent_dt_utc:
                continue
            text = row.get("Text") or ""
            text_norm = normalize_text(text)
            outgoing_rows.append(CsvRow(
                chat_session=row.get("Chat Session") or "",
                sent_dt_utc=sent_dt_utc,
                type=type_val,
                text_norm=text_norm,
                raw_text=text,
            ))
            time_index.append(sent_dt_utc)
    # Sort by time (keep parallel order)
    combined = sorted(zip(time_index, outgoing_rows), key=lambda x: x[0])
    time_index = [t for t, _ in combined]
    outgoing_rows = [r for _, r in combined]
    return outgoing_rows, time_index


def bisect_time_window(times: List[datetime], target: datetime, window: timedelta) -> Tuple[int, int]:
    """Return start and end indices (slice) of times within [target-window, target+window]."""
    import bisect
    start_dt = target - window
    end_dt = target + window
    left = bisect.bisect_left(times, start_dt)
    right = bisect.bisect_right(times, end_dt)
    return left, right


def find_match(
    outgoing_rows: List[CsvRow],
    time_index: List[datetime],
    jsonl_text: str,
    jsonl_dt_utc: Optional[datetime],
) -> Tuple[str, Optional[CsvRow], str]:
    """
    Attempt to match a JSONL whatsapp record to an Outgoing CSV row.
    Returns (decision, matched_row, reason)
      decision: "keep" | "drop_no_match"
    """
    text_norm = normalize_text(jsonl_text)

    # Require a timestamp for matching in strict mode
    if not jsonl_dt_utc:
        return "drop_no_match", None, "missing_jsonl_timestamp"

    # Primary: text+time match within ±5 minutes
    if text_norm:
        left, right = bisect_time_window(time_index, jsonl_dt_utc, TEXT_TIME_WINDOW)
        candidates = outgoing_rows[left:right]
        text_matches = [r for r in candidates if r.text_norm == text_norm]
        if text_matches:
            # Choose closest in time among text matches
            best = min(text_matches, key=lambda r: abs((r.sent_dt_utc - jsonl_dt_utc).total_seconds()))
            return "keep", best, "text_time_match"

    # Fallback: time-only unique match within ±90s
    left2, right2 = bisect_time_window(time_index, jsonl_dt_utc, TIME_ONLY_WINDOW)
    cand2 = outgoing_rows[left2:right2]
    if len(cand2) == 1:
        return "keep", cand2[0], "time_only_unique"

    return "drop_no_match", None, "no_reliable_match"


@dataclass
class AuditRow:
    dataset: str  # training/validation
    chunk_id: str
    timestamp: str
    output: str
    match_reason: str
    csv_chat_session: str = ""
    csv_sent_date_utc: str = ""
    csv_type: str = ""
    csv_text: str = ""


def process_jsonl(
    jsonl_path: Path,
    outgoing_rows: List[CsvRow],
    time_index: List[datetime],
    out_jsonl_path: Path,
    dataset_name: str,
) -> Tuple[int, int, int, List[AuditRow], List[AuditRow], List[AuditRow]]:
    kept = 0
    dropped_matched_non_outgoing = 0  # For completeness; shouldn't happen since we pre-filter outgoing
    dropped_no_match = 0

    kept_rows: List[AuditRow] = []
    dropped_rows: List[AuditRow] = []
    no_match_rows: List[AuditRow] = []

    with jsonl_path.open("r", encoding="utf-8") as fin, out_jsonl_path.open("w", encoding="utf-8") as fout:
        for line in fin:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                # Skip malformed lines but log as no-match
                dropped_no_match += 1
                no_match_rows.append(AuditRow(dataset_name, "", "", line[:200], "json_decode_error"))
                continue

            source = (obj.get("source") or "").lower()
            if source != "whatsapp":
                # Keep non-whatsapp as-is
                fout.write(json.dumps(obj, ensure_ascii=False) + "\n")
                kept += 1
                kept_rows.append(AuditRow(dataset_name, obj.get("chunk_id", ""), obj.get("timestamp", ""), obj.get("output", ""), "non_whatsapp"))
                continue

            ts = obj.get("timestamp") or ""
            dt_utc = parse_jsonl_timestamp(ts)
            output_text = obj.get("output") or ""

            decision, matched_row, reason = find_match(outgoing_rows, time_index, output_text, dt_utc)

            if decision == "keep" and matched_row is not None:
                # Keep this whatsapp record
                fout.write(json.dumps(obj, ensure_ascii=False) + "\n")
                kept += 1
                kept_rows.append(AuditRow(
                    dataset_name,
                    obj.get("chunk_id", ""),
                    ts,
                    output_text,
                    reason,
                    csv_chat_session=matched_row.chat_session,
                    csv_sent_date_utc=matched_row.sent_dt_utc.isoformat(),
                    csv_type=matched_row.type,
                    csv_text=matched_row.raw_text,
                ))
            elif decision == "drop_no_match":
                dropped_no_match += 1
                no_match_rows.append(AuditRow(
                    dataset_name,
                    obj.get("chunk_id", ""),
                    ts,
                    output_text,
                    reason,
                ))
            else:
                # Should not hit because we only loaded outgoing rows
                dropped_matched_non_outgoing += 1
                dropped_rows.append(AuditRow(
                    dataset_name,
                    obj.get("chunk_id", ""),
                    ts,
                    output_text,
                    "matched_non_outgoing",
                ))

    return kept, dropped_matched_non_outgoing, dropped_no_match, kept_rows, dropped_rows, no_match_rows


def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)


def write_audit_xlsx(
    xlsx_path: Path,
    summary: Dict[str, int],
    kept_rows: List[AuditRow],
    dropped_rows: List[AuditRow],
    no_match_rows: List[AuditRow],
):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "summary"
    ws_sum.append(["metric", "count"])
    for k, v in summary.items():
        ws_sum.append([k, v])
    autosize_columns(ws_sum)

    def add_sheet(title: str, rows: List[AuditRow]):
        ws = wb.create_sheet(title)
        ws.append([
            "dataset", "chunk_id", "timestamp", "output", "match_reason",
            "csv_chat_session", "csv_sent_date_utc", "csv_type", "csv_text",
        ])
        for r in rows:
            ws.append([
                r.dataset, r.chunk_id, r.timestamp, r.output, r.match_reason,
                r.csv_chat_session, r.csv_sent_date_utc, r.csv_type, r.csv_text,
            ])
        autosize_columns(ws)

    add_sheet("kept_matches", kept_rows)
    add_sheet("dropped_matches", dropped_rows)
    add_sheet("no_match", no_match_rows)

    wb.save(xlsx_path)


def main():
    # Load outgoing-only CSV rows and build time index
    outgoing_rows, time_index = load_outgoing_csv(CSV_PATH)

    # Process training
    tr_kept, tr_drop_non_out, tr_drop_no, tr_kept_rows, tr_dropped_rows, tr_no_match_rows = process_jsonl(
        TRAIN_JSONL, outgoing_rows, time_index, OUT_TRAIN_JSONL, dataset_name="training"
    )

    # Process validation
    va_kept, va_drop_non_out, va_drop_no, va_kept_rows, va_dropped_rows, va_no_match_rows = process_jsonl(
        VAL_JSONL, outgoing_rows, time_index, OUT_VAL_JSONL, dataset_name="validation"
    )

    summary = {
        "training_kept": tr_kept,
        "training_dropped_matched_non_outgoing": tr_drop_non_out,
        "training_dropped_no_match": tr_drop_no,
        "validation_kept": va_kept,
        "validation_dropped_matched_non_outgoing": va_drop_non_out,
        "validation_dropped_no_match": va_drop_no,
        "total_kept": tr_kept + va_kept,
        "total_dropped_no_match": tr_drop_no + va_drop_no,
    }

    # Write audit workbook
    kept_all = tr_kept_rows + va_kept_rows
    dropped_all = tr_dropped_rows + va_dropped_rows
    no_match_all = tr_no_match_rows + va_no_match_rows
    write_audit_xlsx(AUDIT_XLSX, summary, kept_all, dropped_all, no_match_all)

    # Print concise counts for terminal
    print("Filtering complete.")
    for k, v in summary.items():
        print(f"{k}: {v}")


if __name__ == "__main__":
    main()
